from os import path
import csv

import openpyxl
import xlrd


class IndicatorsProcessor(object):

    def __init__(self, file_paths):
        # Set file paths
        self.file_paths = file_paths
        self.input_files = {
            'econ_freedom': path.abspath(file_paths['input_dir']+'/'+'economic-freedom-of-the-world-2015-dataset.xlsx'),
            'doing_business': path.abspath(file_paths['input_dir']+'/'+'Data_Extract_From_Doing_Business.xlsx'),
            'econ_heritage': path.abspath(file_paths['input_dir']+'/'+'economic_freedom_heritage-ok.xlsx'),
            'cpi_db': path.abspath(file_paths['input_dir']+'/'+'cpi_okfn-ok.csv'),
            'fin_open': path.abspath(file_paths['input_dir']+'/'+'kaopen_2013.xls'),
            'polcon': path.abspath(file_paths['input_dir']+'/'+'polcon2012-ok.xls'),
            'shadow': path.abspath(file_paths['input_dir']+'/'+'shadow_eco_knoema.xlsx'),
            'ulc': path.abspath(file_paths['input_dir']+'/'+'ulc_e2003.xls'),
        }
        # Open input file
        out_path = path.abspath(file_paths['output_dir']+'/'+'data_input.xlsx')
        self.out_workbook = openpyxl.load_workbook(filename=out_path)
        # Set source start positions (columns)
        self.source_pos = {
            'cpi': 'E', 'fin_open': 'F', 'global': 'G', 'polcon': 'K', 'shadow': 'M', 'ulc': 'N',
            'econ_freedom': 'O', 'econ_heritage': 'R', 'do_biz': 'AC'
        }

    def write_years_and_countries(self, from_year, to_year):
        """
        Write years and countries to out worksheet data.
        """
        # Get country codes (and names) from out workbook.
        countries_ws = self.out_workbook.get_sheet_by_name('countries')
        countries_data = []
        for row in countries_ws.iter_rows('B2:D252'):
            name = row[0].value.strip()
            code2 = row[1].value.strip()
            code3 = row[2].value.strip()
            countries_data.append([code3, name, code2])
        # Write data
        data_ws = self.out_workbook.get_sheet_by_name('data')
        write_row_n = 3
        for country in countries_data:
            for year in range(from_year, to_year+1):
                data_ws.cell(row=write_row_n, column=1).value = year
                data_ws.cell(row=write_row_n, column=2).value = country[0]
                data_ws.cell(row=write_row_n, column=3).value = country[2]
                data_ws.cell(row=write_row_n, column=4).value = country[1]
                write_row_n += 1
        # Save file
        self.out_workbook.save(self.file_paths['output_dir']+'/'+'data_input.xlsx')

    def write_workbook_to_file(self):
        self.out_workbook.save(self.file_paths['output_dir']+'/'+'data_output.xlsx')

    def write_econ_freedom(self):
        """
        Economic freedom of the world:
        -> Business regulations, Freedom to trade internationaly, Top Marginal Tax Rate
        """
        # Open input
        in_wb = openpyxl.load_workbook(self.input_files['econ_freedom'])
        in_sheet = in_wb.get_sheet_by_name('Unadjusted Data')
        out_sheet = self.out_workbook.get_sheet_by_name('data')
        # Prepare input and output rows
        in_rows = in_sheet.rows[5:]
        out_rows = out_sheet.rows[2:]
        last_out_row_n = 0
        # Get numeric column index for the first output item.
        out_first_col_idx = openpyxl.utils.column_index_from_string(self.source_pos['econ_freedom'])
        # Read all input rows.
        for row_in in in_rows:
            # Check if the column is not empty - the end.
            if not row_in[1].value:
                break
            # Read values.
            year_in = int(row_in[1].value)
            code_in = row_in[2].value.strip()
            print('IN %d: %s') % (year_in, code_in)
            # Get scores
            br_col_idx = openpyxl.utils.column_index_from_string('BQ')
            br_value = row_in[br_col_idx - 1].value
            freetr_idx = openpyxl.utils.column_index_from_string('AY')
            freetr_value = row_in[freetr_idx - 1].value
            topmar_idx = openpyxl.utils.column_index_from_string('P')
            topmar_value = row_in[topmar_idx - 1].value
            # Write scores to the correct output row.
            for row_n, row_out in enumerate(out_rows[last_out_row_n:], start=last_out_row_n):
                year_out = int(row_out[0].value)
                code_out = row_out[1].value.strip()
                #print('%d OUT %d: %s') % (i, year_out, code_out)
                if year_out == year_in and code_out == code_in:
                    row_out[out_first_col_idx - 1].value = br_value
                    row_out[out_first_col_idx].value = freetr_value
                    row_out[out_first_col_idx + 1].value = topmar_value
                    last_out_row_n = 0
                    break
                else:
                    continue
        # end

    def write_econ_heritage(self):
        """
        Index of economic freedom (Heritage.org)
        Write all fields in order of appereance.
        """
        # Open input
        in_wb = openpyxl.load_workbook(self.input_files['econ_heritage'])
        in_sheet = in_wb.active
        out_sheet = self.out_workbook.get_sheet_by_name('data')
        # Prepare input and output rows
        in_rows = in_sheet.rows[1:]
        out_rows = out_sheet.rows[2:]
        last_out_row_n = 0
        # Get numeric column index for the first output item.
        out_first_col_idx = openpyxl.utils.column_index_from_string(self.source_pos['econ_heritage']) - 1
        # Read all input rows.
        for row_in in in_rows:
            # Check if the column is not empty - the end.
            if not row_in[0].value:
                break
            # Read values.
            year_in = int(row_in[1].value)
            cname_in = row_in[0].value.strip().lower()
            print('IN %d: %s') % (year_in, cname_in)
            # Get scores
            out_scores = [float(row_in[x].value) if row_in[x].value != 'N/A' else '' for x in range(2, 13)]
            # Write scores to the correct output row.
            for row_n, row_out in enumerate(out_rows[last_out_row_n:], start=last_out_row_n):
                year_out = int(row_out[0].value)
                cname_out = row_out[3].value.strip().lower()
                #print('%d OUT %d: %s') % (row_n, year_out, cname_out)
                if year_out == year_in and cname_out == cname_in:
                    for i_s, score in enumerate(out_scores, out_first_col_idx):
                        row_out[i_s].value = score
                    last_out_row_n = 0
                    break
                else:
                    continue
        # end

    def write_cpi(self):
        """
        Corruption Perceptions Index
        """
        # Prepare stuff
        in_file = open(self.input_files['cpi_db'], 'rb')
        cpi_reader = csv.reader(in_file)
        out_sheet = self.out_workbook.get_sheet_by_name('data')
        out_rows = out_sheet.rows[2:]
        last_out_row_n = 0
        out_first_col_idx = openpyxl.utils.column_index_from_string(self.source_pos['cpi']) - 1
        # Get the first (header) line.
        header_line = cpi_reader.next()
        first_year, last_year = int(header_line[1]), int(header_line[-1])
        # Read countries (lines) from file.
        for row in cpi_reader:
            cname_in = row[0].strip().lower()
            print('IN %s') % cname_in
            # Save yearly values (1998-2014).
            yearly_values = [float(row[x]) if row[x] != 'NA' else '' for x in range(1, len(header_line))]
            # Find the first position of the country in the out workbook.
            for row_n, row_out in enumerate(out_rows[last_out_row_n:], start=last_out_row_n):
                year_out = int(row_out[0].value)
                cname_out = row_out[3].value.strip().lower()
                #print('%d OUT %d: %s') % (row_n, year_out, cname_out)
                if year_out == first_year and cname_out == cname_in:
                    last_out_row_n = row_n
                    # Write values for all subsequent years.
                    for y_value in yearly_values:
                        row_out[out_first_col_idx].value = y_value
                        last_out_row_n += 1
                        row_out = out_rows[last_out_row_n]
                    # reset - for sure
                    last_out_row_n = 0
                    break
                else:
                    continue
        # end

    def write_finopen(self):
        """
        Financial Openness
        """
        # Open input
        in_wb = xlrd.open_workbook(self.input_files['fin_open'])
        in_sheet = in_wb.sheet_by_index(0)
        out_sheet = self.out_workbook.get_sheet_by_name('data')
        # Prepare input and output rows
        out_rows = out_sheet.rows[2:]
        last_out_row_n = 0
        # Get numeric column index for the first output item.
        out_first_col_idx = openpyxl.utils.column_index_from_string(self.source_pos['fin_open']) - 1
        # Read all input rows.
        for x in range(1, in_sheet.nrows):
            row_in = in_sheet.row(x)
            # Read values.
            year_in = int(row_in[3].value)
            code_in = row_in[1].value.strip()
            print('IN %d: %s') % (year_in, code_in)
            # Get the score.
            ko_value = row_in[4].value
            # Write score to the correct output row.
            for row_n, row_out in enumerate(out_rows[last_out_row_n:], start=last_out_row_n):
                year_out = int(row_out[0].value)
                code_out = row_out[1].value.strip()
                #print('%d OUT %d: %s') % (row_n, year_out, code_out)
                if year_out == year_in and code_out == code_in:
                    row_out[out_first_col_idx].value = ko_value
                    last_out_row_n = 0  # Or set to row_n, if the speed is important (1/4 time).
                    break
                else:
                    continue
        # end

    def write_global(self):
        """
        KOF globalization index
        Globalization Index, Economic Globalization, Social Globalization, Political Globalization
        """
        indices = ['index', 'economic', 'social', 'political']
        for i_n, i_name in enumerate(indices):
            # Prepare stuff
            in_file = open(path.abspath(self.file_paths['input_dir']+'/global_'+i_name+'.csv'), 'rb')
            csv_reader = csv.reader(in_file)
            out_sheet = self.out_workbook.get_sheet_by_name('data')
            out_rows = out_sheet.rows[2:]
            last_out_row_n = 0
            out_col_idx = openpyxl.utils.column_index_from_string(self.source_pos['global']) - 1 + i_n
            # Get the first (header) line.
            header_line = csv_reader.next()
            first_year = int(header_line[2])
            # Read countries (lines) from file.
            for row in csv_reader:
                ccode_in = row[0].strip()
                print('IN %s') % ccode_in
                # Save yearly values (1970-2012).
                yearly_values = [float(row[x]) if row[x] != '.' else '' for x in range(2, len(header_line))]
                # Find the first position of the country in the out workbook.
                for row_n, row_out in enumerate(out_rows[last_out_row_n:], start=last_out_row_n):
                    year_out = int(row_out[0].value)
                    ccode_out = row_out[1].value.strip()
                    #print('%d OUT %d: %s') % (row_n, year_out, cname_out)
                    if year_out == first_year and ccode_out == ccode_in:
                        last_out_row_n = row_n
                        # Write values for all subsequent years.
                        for y_value in yearly_values:
                            row_out[out_col_idx].value = y_value
                            last_out_row_n += 1
                            row_out = out_rows[last_out_row_n]
                        last_out_row_n = 0
                        break
                    else:
                        continue
            # end

    def write_polcon(self):
        """
        Political Constrains Index (III and V)
        """
        # Open input
        in_wb = xlrd.open_workbook(self.input_files['polcon'])
        in_sheet = in_wb.sheet_by_index(1)
        out_sheet = self.out_workbook.get_sheet_by_name('data')
        # Prepare input and output rows
        out_rows = out_sheet.rows[2:]
        last_out_row_n = 0
        # Get numeric column index for the first output item.
        out_first_col_idx = openpyxl.utils.column_index_from_string(self.source_pos['polcon']) - 1
        # Read all input rows.
        for x in range(1, in_sheet.nrows):
            row_in = in_sheet.row(x)
            # Check if the column is not empty - the end.
            if not row_in[0].value:
                break
            # Read values.
            year_in = int(row_in[6].value)
            cname_in = row_in[2].value.strip().lower()
            print('IN %d: %s') % (year_in, cname_in)
            # Get the score.
            iii_value = row_in[7].value
            v_value = row_in[8].value
            # Write score to the correct output row.
            for row_n, row_out in enumerate(out_rows[last_out_row_n:], start=last_out_row_n):
                year_out = int(row_out[0].value)
                cname_out = row_out[3].value.strip().lower()
                #print('%d OUT %d: %s') % (row_n, year_out, cname_out)
                if year_out == year_in and cname_out == cname_in:
                    row_out[out_first_col_idx].value = iii_value
                    row_out[out_first_col_idx + 1].value = v_value
                    last_out_row_n = 0
                    break
                else:
                    continue
        # end

    def write_shadow(self):
        """
        Shadow Economy Size: Size as Percentage of Official GDP
        """
        # Open input
        in_wb = openpyxl.load_workbook(self.input_files['shadow'])
        in_sheet = in_wb.active
        out_sheet = self.out_workbook.get_sheet_by_name('data')
        # Prepare input and output rows
        in_rows = in_sheet.rows[1:]
        out_rows = out_sheet.rows[2:]
        last_out_row_n = 0
        # Get the first (header) line.
        header_line = in_sheet.rows[0]
        first_year = int(header_line[8].value)
        # Get numeric column index for the output item.
        out_col_idx = openpyxl.utils.column_index_from_string(self.source_pos['shadow']) - 1
        # Read all input rows.
        for row_in in in_rows:
            # Check if the column is not empty - the end.
            if not row_in[0].value:
                break
            # Read values.
            ccode_in = row_in[2].value.strip()
            print('IN %s') % ccode_in
            # Save values for all years.
            yearly_values = [float(row_in[x].value) if row_in[x].value != '' else '' for x in range(8, len(header_line))]
            # Write values to the correct output row.
            for row_n, row_out in enumerate(out_rows[last_out_row_n:], start=last_out_row_n):
                year_out = int(row_out[0].value)
                ccode_out = row_out[2].value.strip()
                #print('%d OUT %d: %s') % (row_n, year_out, cname_out)
                if year_out == first_year and ccode_out == ccode_in:
                    last_out_row_n = row_n
                    # Write values for all subsequent years.
                    for y_value in yearly_values:
                        row_out[out_col_idx].value = y_value
                        last_out_row_n += 1
                        row_out = out_rows[last_out_row_n]
                    last_out_row_n = 0
                    break
                else:
                    continue
        # end

    def write_dobiz(self):
        """
        Doing Business database
        50+ indicators
        """
        # Open input
        in_wb = xlrd.open_workbook(self.input_files['doing_business'])
        in_sheet = in_wb.sheet_by_index(0)
        out_sheet = self.out_workbook.get_sheet_by_name('data')
        # Prepare input and output rows
        out_rows = out_sheet.rows[2:]
        last_out_row_n = 0
        # Get the first (header) line.
        header_line = in_sheet.row(0)
        first_year = int(header_line[4].value[0:4])
        # Get numeric column index for the first output item.
        out_first_col_idx = openpyxl.utils.column_index_from_string(self.source_pos['do_biz']) - 1
        # Read all input rows.
        for r_i in range(1, in_sheet.nrows):
            # Get current row.
            row_in = in_sheet.row(r_i)
            # Read country code.
            ccode_in = row_in[1].value.strip()
            print('IN %s') % ccode_in
            if r_i == 1:
                current_code = ccode_in
                indicators_yearly_data = []
            # Save values for all years.
            yearly_values = [float(row_in[x].value)
                             if (row_in[x].value != '..' and row_in[x].value != '') else ''
                             for x in range(4, len(header_line))]
            # Check if I still read code for the same country.
            if current_code == ccode_in:
                # YES - Add yearly values to the list and continue reading.
                indicators_yearly_data.append(yearly_values)
                current_code == ccode_in
                continue
            else:
                # NO
                # Process yearly data into row data: one year and values for all indicators.
                country_row_data = {}
                available_years = [int(x.value[0:4]) for x in header_line[4:]]
                for year in available_years:
                    country_row_data[year] = []
                for indicator_data in indicators_yearly_data:
                    for i_year, year in enumerate(available_years):
                        country_row_data[year].append(indicator_data[i_year])
                # Write data
                for row_n, row_out in enumerate(out_rows[last_out_row_n:], start=last_out_row_n):
                    year_out = int(row_out[0].value)
                    ccode_out = row_out[1].value.strip()
                    #print('%d OUT %d: %s') % (row_n, year_out, cname_out)
                    # Find the first occurence of the country in the workbook.
                    if year_out == first_year and ccode_out == current_code:
                        last_out_row_n = row_n
                        # Write all data year by year.
                        for year, y_data in country_row_data.items():
                            # For current year, write values for all indicators.
                            for i_s, i_value in enumerate(y_data, out_first_col_idx):
                                row_out[i_s].value = i_value
                            # Go to the next line (year)
                            last_out_row_n += 1
                            row_out = out_rows[last_out_row_n]
                        # reset stuff
                        last_out_row_n = 0
                        break
                    else:
                        continue
                # Reset data
                current_code = ccode_in
                indicators_yearly_data = [yearly_values]
                # Check if we should end.
                if r_i == 12509:
                    break
        # end

    def write_ulc(self):
        """
        Unit Labour Costs (OECD)
        """
        # Open input
        in_wb = xlrd.open_workbook(self.input_files['ulc'])
        in_sheet = in_wb.sheet_by_index(0)
        out_sheet = self.out_workbook.get_sheet_by_name('data')
        # Prepare input and output rows
        out_rows = out_sheet.rows[2:]
        last_out_row_n = 0
        # Get the first (header) line.
        header_line = in_sheet.row(6)
        first_year = int(header_line[3].value)
        # Get numeric column index for the first output item.
        out_col_idx = openpyxl.utils.column_index_from_string(self.source_pos['ulc']) - 1
        # Read all input rows.
        for i_r in range(8, in_sheet.nrows):
            row_in = in_sheet.row(i_r)
            # The end.
            if i_r == 42:
                break
            # Special reading for rows 40 and 41.
            if i_r == 40 or i_r == 41:
                cname_in = row_in[1].value.strip()
            else:
                cname_in = row_in[0].value.strip()
            print('IN %d: %s') % (i_r, cname_in)
            # Save values for all years.
            yearly_values = [float(row_in[x].value) if row_in[x].value != '..' else '' for x in range(3, len(header_line))]
            # Write values to the correct output row.
            for row_n, row_out in enumerate(out_rows[last_out_row_n:], start=last_out_row_n):
                year_out = int(row_out[0].value)
                cname_out = row_out[3].value.strip()
                #print('%d OUT %d: %s') % (row_n, year_out, cname_out)
                if year_out == first_year and cname_out == cname_in:
                    last_out_row_n = row_n
                    # Write values for all subsequent years.
                    for y_value in yearly_values:
                        row_out[out_col_idx].value = y_value
                        last_out_row_n += 1
                        row_out = out_rows[last_out_row_n]
                    last_out_row_n = 0
                    break
                else:
                    continue
        # end
