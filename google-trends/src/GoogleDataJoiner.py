from GoogleDataGetter import GoogleDataGetter
import openpyxl
import os

import codecs


class GoogleDataJoiner(object):
    """
    Used to join google data files to one excel file.
    """

    def __init__(self, file_paths):
        # Set file paths
        self.file_paths = file_paths
        # Create an output workbook
        self.out_workbook = openpyxl.Workbook()
        self.out_workbook.remove_sheet(self.out_workbook.active)    # remove defaultly created sheet
        # Read country codes.
        self.country_codes = GoogleDataGetter.get_country_codes(file_paths['country_codes_file'])

    def process_all_countries(self):
        for c_code in self.country_codes.keys():
            print('====PROCESSING %s====' % c_code)
            self.process_country(c_code)
        # Save workbook to file
        out_filename = os.path.abspath(self.file_paths['output_dir']+'/google_data_auto.xlsx')
        self.out_workbook.save(out_filename)

    def process_country(self, c_code):
        # Get country directory.
        country_dir = self.file_paths['gdata_dir']+'/'+c_code
        country_sheet = self.out_workbook.create_sheet(c_code)
        # Browse through all files in the folder.
        for file_name in sorted(os.listdir(country_dir)):
            # Read term data from file.
            term_data = self._read_data_from_gfile(file_name, country_dir)
            # Save term data to sheet.
            self._write_term_data_to_sheet(country_sheet, term_data)
        # OK

    def _read_data_from_gfile(self, file_name, dir_path):
        # Open file
        gfile = codecs.open(dir_path+'/'+file_name, 'r', 'utf-8')
        # Get term ID
        term_id = file_name.split('_')[1].replace('.csv', '')
        # Get search term
        term_text = gfile.readline().split(':')[1].strip()
        # Skip 4 lines
        for i in range(0, 4):
            gfile.readline()
        # Save all data until empty line.
        term_values = []
        for line in gfile:
            if not line.strip():
                break
            term_values.append(line.strip())
        # Close file and return data.
        gfile.close()
        return [term_id, term_text, term_values]

    def _write_term_data_to_sheet(self, country_sheet, term_data):
        # Get items from list.
        term_id = int(term_data[0])
        term_text = term_data[1]
        term_values = term_data[2]
        # Write term ID and term text.
        country_sheet.cell(column=term_id, row=1).value = term_id
        country_sheet.cell(column=term_id, row=2).value = term_text
        # Are there any time values to write?
        if not term_values:
            return None
        # If yes, write them all!
        for row_idx, t_value in enumerate(term_values, 3):
            country_sheet.cell(column=term_id, row=row_idx).value = t_value
