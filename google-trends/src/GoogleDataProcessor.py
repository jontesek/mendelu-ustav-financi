from os import path
import re

import openpyxl


class GoogleDataProcessor(object):

    def __init__(self, file_paths, default_lang, last_year):
        # Set file paths
        self.file_paths = file_paths
        wb_filepath = path.abspath(file_paths['input_file'])
        # Read an input workbook
        self.in_workbook = openpyxl.load_workbook(filename=wb_filepath)
        self.countries = self.in_workbook.get_sheet_names()
        # Create an output workbook
        self.out_workbook = openpyxl.Workbook()
        self.out_workbook.remove_sheet(self.out_workbook.active)    # remove defaultly created sheet
        # Regexp
        self.reg_week = re.compile('^(\d{4}-\d{2}-\d{2}) - (\d{4}-\d{2}-\d{2}),(\d+)')
        self.reg_month = re.compile('^(\d{4}-\d{2}),(\d+)')
        # Set default language for words in one sheet
        self.default_lang = default_lang
        # Last year for processing
        self.last_year = last_year

    def process_all_countries_to_one_sheet(self):
        # Create main and only sheet.
        one_sheet = self.out_workbook.create_sheet('together')
        # Write sheet header
        one_sheet.cell('A1').value = 'period'
        one_sheet.cell('B1').value = 'country'
        # Write words in default language
        for idx, col_word in enumerate(self.in_workbook[self.default_lang].columns):
            if not col_word.value:
                break
            one_sheet.cell(column=idx+3,row=1).value = col_word[1].value
        # Process all countries and their data.
        for country in self.in_workbook.get_sheet_names():
            pass
        # Save file
        out_filename = path.abspath(self.file_paths['output_dir']+'/total_period_data.xlsx')
        self.out_workbook.save(out_filename)

    def process_all_countries(self):
        for country_code in self.in_workbook.get_sheet_names():
            print('====PROCESSING %s====' % country_code)
            self.process_country(country_code)
            if country_code == 'CZ':
                pass#self.process_country(country_code)
            else:
                continue
        # Save workbook to file
        out_filename = path.abspath(self.file_paths['output_dir']+'/period_data.xlsx')
        self.out_workbook.save(out_filename)


    def process_country(self, country_code):
        # Read the country sheet.
        in_sheet = self.in_workbook[country_code]
        # Create a new empty sheet.
        country_sheet = self.out_workbook.create_sheet(country_code)
        # Browse all columns (search terms).
        country_data = []
        for col in in_sheet.columns:
            # Check if an empty column was reached.
            if not col[0].value:
                break    # stop the loop
            country_data.append(self._process_search_word(col))
            if col[0].value > 1:
                pass#break
        # Write data to the sheet.
        self._write_country_data(country_sheet, country_data)


    def _process_search_word(self, column):
        # Get word info
        word_id = column[0].value
        word_string = column[1].value
        # Check if the first item is empty - data not present.
        if not column[2].value:
            #print('===empty===')
            return [word_id, word_string, False]
        # Check if the values have monthly or weekly frequency.
        if self.reg_week.match(column[2].value):
            #print('===week===')
            word_data = self._process_weekly_values(3, column[2:])
        else:
            #print('===month===')
            word_data = self._process_monthly_values(3, column[2:])
        # result
        # print word_data
        return [word_id, word_string, word_data]

    def _write_country_data(self, country_sheet, country_data):
        # Write first column header (period names)
        country_sheet.cell('A1').value = 0
        country_sheet.cell('A2').value = 'period'
        # Write all search words and associated data
        for col_idx, (word_id, word_string, word_data) in enumerate(country_data):
            # Set column number
            col_num = col_idx + 2
            # Write word ID and word string.
            country_sheet.cell(column=col_num,row=1).value = word_id
            country_sheet.cell(column=col_num,row=2).value = word_string
            # Check if there are any data.
            if not word_data:
                continue    # If not, skip the word.
            # If yes, write periods data.
            for row_idx, (period_name, avg_num) in enumerate(word_data):
                # If it is the first processed column, write period name to the first column.
                if col_idx == 0:
                    country_sheet.cell(column=1,row=row_idx+3).value = period_name
                # Write period average
                country_sheet.cell(column=col_num,row=row_idx+3).value = avg_num


    def _process_monthly_values(self, months_period, items):
        # Prepare variables
        word_data = []
        period = {'year': None, 'sum': 0.0, 'number': 1}
        # Browse all rows
        for index, item in enumerate(items):
            # Check if an empty row was reached.
            if not item.value:
                break    # stop the loop
            # Get values
            #print item.value
            result = self.reg_month.match(item.value.strip())
            result_count = float(result.group(2))
            result_year = int(result.group(1)[0:4])
            # Check if the script should stop processing values.
            if result_year > self.last_year:
                break
            # If not, continue.
            period['sum'] = result_count + period['sum']
            period['year'] = result_year
            #print '%d: %d' % (int(index+1) % months_period, result_count)
            # Check if the current period should end (i.e. there are 3 months in a quarter of a year).
            if int(index+1) % months_period == 0:
                #print period
                # If yes, count the average score of the word and save it with the period ID.
                period_avg = period['sum'] / months_period
                period_name = '%sq%s' % (period['year'], period['number'])
                word_data.append((period_name, period_avg))
                # Reset the period
                period = {'year': None, 'sum': 0, 'number': period['number']+1}
                # If the new year is on the plan, reset period number.
                if period['number'] > 4:
                    period['number'] = 1
        # return data
        return word_data


    def _process_weekly_values(self, months_period, items):
        # Prepare variables
        one_month = []
        all_months = []
        # Browse all rows and insert all months.
        for index, item in enumerate(items):
            # Check if an empty row was reached.
            if not item.value:
                break    # stop the loop
            # Get values
            #print item.value
            reg_week = re.compile('^(\d{4})-(\d{2})-\d{2} - (\d{4})-(\d{2})-\d{2},(\d+)')
            result = reg_week.match(item.value.strip())
            result_count = float(result.group(5))
            result_dates = {
                'year_from': int(result.group(1)),
                'month_from': int(result.group(2)),
                'year_to': int(result.group(3)),
                'month_to': int(result.group(4))
            }
            if index == 0:
                one_month.append((index+3, result_dates['month_from'], result_count, result_dates['year_from']))
            else:
                # Check if start month is the same as previous end month.
                if result_dates['month_from'] == one_month[len(one_month)-1][1]:
                    # Add week to the month
                    one_month.append((index+3, result_dates['month_from'], result_count, result_dates['year_from']))
                else:
                    # Count month average and add month to the list.
                    all_months.append(self._create_month_from_weeks(one_month, one_month[-1][3]))
                    #print one_month
                    #print all_months[-1]
                    # Reset month and add the week as the new month
                    one_month = []
                    one_month.append((index+3, result_dates['month_from'], result_count, result_dates['year_from']))
                        # Check if the script should stop processing values.
            if result_dates['year_from'] > self.last_year:
                break
        # Count periods average from all months.
        return self._create_periods_from_months(months_period,all_months)

    def _create_month_from_weeks(self, weeks_list, year_from):
        count_sum = 0.0
        for week in weeks_list:
            count_sum += week[2]
        avg_count = count_sum / len(weeks_list)
        return (year_from, avg_count)

    def _create_periods_from_months(self, months_period, months_list):
        # Prepare variables
        word_data = []
        period = {'year': None, 'sum': 0, 'number': 1}
        # Browse all months and create periods.
        for idx, (year_from, month_count) in enumerate(months_list):
            #print (year_from, month_count)
            period['sum'] = month_count + period['sum']
            period['year'] = year_from
            #print '%d: %d' % (int(idx+1) % months_period, month_count)
            # Check if the current period should end (i.e. there are 3 months in a quarter of a year).
            if int(idx+1) % months_period == 0:
                # If yes, count the average score of the word and save it with the period ID.
                period_avg = period['sum'] / months_period
                period_name = '%sq%s' % (period['year'], period['number'])
                #print '%d/%d' % (period['sum'], months_period),
                #print (period_name, period_avg)
                word_data.append((period_name, period_avg))
                # Reset the period
                period = {'year': None, 'sum': 0, 'number': period['number']+1}
                # If the new year is on the plan, reset period number.
                if period['number'] > 4:
                    period['number'] = 1
        return word_data
